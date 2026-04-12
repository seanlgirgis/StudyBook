#!/usr/bin/env python3
"""
Search coding_challenges/index.xlsx by substring (grep-like).
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INDEX_PATH = ROOT / "index.xlsx"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Search coding_challenges/index.xlsx by substring")
    parser.add_argument("needle", help="Substring to search for (example: 48)")
    parser.add_argument("--index", type=Path, default=DEFAULT_INDEX_PATH, help="Path to .xlsx index file")
    parser.add_argument("--case-sensitive", action="store_true", help="Use case-sensitive search")
    parser.add_argument("--limit", type=int, default=50, help="Max rows to print")
    return parser.parse_args()


def normalize(text: str, case_sensitive: bool) -> str:
    return text if case_sensitive else text.lower()


def main() -> int:
    args = parse_args()
    index_path = args.index.resolve()

    if not index_path.exists():
        raise SystemExit(f"Index file does not exist: {index_path}")

    wb = load_workbook(index_path, read_only=True, data_only=True)
    ws = wb["index"] if "index" in wb.sheetnames else wb.active

    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_cells:
        print("No headers found.")
        return 0

    headers = [str(h).strip() if h is not None else "" for h in header_cells]
    needle = normalize(args.needle, args.case_sensitive)

    matches: list[str] = []
    total = 0
    for excel_row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        parts = []
        for i, value in enumerate(row):
            key = headers[i] if i < len(headers) and headers[i] else f"col_{i+1}"
            text = "" if value is None else str(value)
            parts.append((key, text))

        searchable = " | ".join(text for _, text in parts)
        if needle in normalize(searchable, args.case_sensitive):
            total += 1
            rendered = " | ".join(f"{k}={v}" for k, v in parts)
            matches.append(f"{excel_row_num}: {rendered}")
            if len(matches) >= args.limit:
                break

    if total == 0:
        print(f"No matches for '{args.needle}' in {index_path}")
        return 0

    print(f"Matches for '{args.needle}' in {index_path} (showing up to {args.limit})")
    for line in matches:
        print(line)
    if total > len(matches):
        print(f"... and at least {total - len(matches)} more in scanned rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
