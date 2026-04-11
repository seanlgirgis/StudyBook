#!/usr/bin/env python3
"""
Refresh coding_challenges/index.xlsx from files under leetcode/by_topic.

Default behavior:
- scans configured solution file extensions
- reads optional metadata header keys when present:
  id, title, tags, source
- falls back to filename/folder-based inference when metadata is absent
- writes deterministic output sorted by path
"""

from __future__ import annotations

import argparse
import csv
import re
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SCAN_ROOT = ROOT / "leetcode" / "by_topic"
DEFAULT_INDEX_PATH = ROOT / "index.xlsx"
DEFAULT_EXTS = {".py"}
HEADERS = ["id", "path", "primary", "tags", "title", "source"]

ID_RE = re.compile(r"^(?:LC|lc)[\-_ ]?(\d{1,5})")
KV_RE = re.compile(r"^\s*([a-zA-Z_][a-zA-Z0-9_]*)\s*:\s*(.+?)\s*$")


@dataclass
class Row:
    row_id: str
    path: str
    primary: str
    tags: str
    title: str
    source: str


def normalize_primary(folder_name: str) -> str:
    return folder_name.strip().lower()


def slug_to_title(slug: str) -> str:
    text = slug.replace("_", " ").replace("-", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text.title() if text else "Untitled"


def normalize_id(file_stem: str) -> str:
    match = ID_RE.match(file_stem)
    if match:
        return f"lc_{int(match.group(1)):04d}"
    return re.sub(r"[^a-z0-9]+", "_", file_stem.lower()).strip("_")


def parse_tags(raw: str) -> list[str]:
    raw = raw.strip()
    if raw.startswith("[") and raw.endswith("]"):
        raw = raw[1:-1]
    parts = [p.strip().strip("'\"") for p in raw.split(",")]
    return [p for p in parts if p]


def read_header_metadata(path: Path) -> dict[str, str]:
    metadata: dict[str, str] = {}

    if path.suffix.lower() not in {".md", ".py"}:
        return metadata

    try:
        lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()
    except OSError:
        return metadata

    for line in lines[:30]:
        if not line.strip():
            continue
        if line.strip().startswith(("#", '"""', "'''", "---")):
            continue
        match = KV_RE.match(line)
        if not match:
            if metadata:
                break
            continue
        key = match.group(1).strip().lower()
        value = match.group(2).strip()
        if key in {"id", "title", "tags", "source"}:
            metadata[key] = value
    return metadata


def build_row(scan_root: Path, file_path: Path) -> Row:
    rel = file_path.relative_to(ROOT).as_posix()
    primary = normalize_primary(file_path.parent.name)
    stem = file_path.stem
    inferred_id = normalize_id(stem)
    inferred_title = slug_to_title(ID_RE.sub("", stem).lstrip("_- ")) if ID_RE.match(stem) else slug_to_title(stem)
    inferred_source = "leetcode"

    meta = read_header_metadata(file_path)

    row_id = meta.get("id", inferred_id).strip()
    title = meta.get("title", inferred_title).strip()
    source = meta.get("source", inferred_source).strip().lower()

    if "tags" in meta:
        tags_list = parse_tags(meta["tags"])
    else:
        tags_list = [primary]

    # ensure primary is always included in tags
    if primary and primary not in tags_list:
        tags_list.insert(0, primary)

    tags = ";".join(dict.fromkeys(tags_list))

    return Row(
        row_id=row_id,
        path=rel,
        primary=primary,
        tags=tags,
        title=title,
        source=source,
    )


def iter_solution_files(scan_root: Path, exts: set[str]) -> Iterable[Path]:
    for path in scan_root.rglob("*"):
        if not path.is_file():
            continue
        if path.suffix.lower() not in exts:
            continue
        yield path


def dedupe_rows(rows: list[Row]) -> list[Row]:
    seen: set[tuple[str, str]] = set()
    out: list[Row] = []
    for row in rows:
        key = (row.row_id, row.path)
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def write_csv(index_path: Path, rows: list[Row]) -> None:
    index_path.parent.mkdir(parents=True, exist_ok=True)
    with index_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        for row in rows:
            writer.writerow([row.row_id, row.path, row.primary, row.tags, row.title, row.source])


def write_xlsx(index_path: Path, rows: list[Row]) -> None:
    index_path.parent.mkdir(parents=True, exist_ok=True)
    if index_path.exists():
        wb = load_workbook(index_path)
        ws = wb["index"] if "index" in wb.sheetnames else wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "index"

    existing_max_row = ws.max_row

    template_styles: dict[int, object] = {}
    if existing_max_row >= 2:
        for col_idx in range(1, len(HEADERS) + 1):
            template_styles[col_idx] = copy(ws.cell(row=2, column=col_idx)._style)

    for col_idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row in enumerate(rows, start=2):
        values = [row.row_id, row.path, row.primary, row.tags, row.title, row.source]
        is_new_row = row_idx > existing_max_row
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if is_new_row and col_idx in template_styles:
                cell._style = copy(template_styles[col_idx])

    last_needed_row = len(rows) + 1
    if existing_max_row > last_needed_row:
        for row_idx in range(last_needed_row + 1, existing_max_row + 1):
            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=col_idx, value=None)

    wb.save(index_path)


def write_index(index_path: Path, rows: list[Row]) -> None:
    suffix = index_path.suffix.lower()
    if suffix == ".csv":
        write_csv(index_path, rows)
        return
    if suffix == ".xlsx":
        write_xlsx(index_path, rows)
        return
    raise SystemExit(f"Unsupported output extension '{index_path.suffix}'. Use .xlsx or .csv.")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Refresh coding_challenges/index.xlsx")
    parser.add_argument("--scan-root", type=Path, default=DEFAULT_SCAN_ROOT, help="Root directory to scan")
    parser.add_argument("--index", type=Path, default=DEFAULT_INDEX_PATH, help="Index output path (.xlsx or .csv)")
    parser.add_argument("--ext", action="append", default=None, help="File extension to include (repeatable)")
    parser.add_argument("--check", action="store_true", help="Check mode: do not write, only report count")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    exts = {e.lower() if e.startswith(".") else f".{e.lower()}" for e in (args.ext or DEFAULT_EXTS)}
    scan_root = args.scan_root.resolve()
    index_path = args.index.resolve()

    if not scan_root.exists():
        raise SystemExit(f"Scan root does not exist: {scan_root}")

    rows = [build_row(scan_root, p) for p in iter_solution_files(scan_root, exts)]
    rows = dedupe_rows(rows)
    rows.sort(key=lambda r: (r.path, r.row_id))

    if args.check:
        print(f"Would index {len(rows)} files into {index_path}")
        return 0

    write_index(index_path, rows)
    print(f"Wrote {len(rows)} rows to {index_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
